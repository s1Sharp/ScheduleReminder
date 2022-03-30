import openpyxl
import logging

from xlsxparser import xlsx_env
from xlsxparser.db_utils import MongodbService, invalid_returned_id

storage = MongodbService.get_instance()

min_len_classwork = 10

INVALID_TEXT = ['\n', '', None]


def update_schedule_db():
    if storage.check_connection() == False:
        logging.error("update schedule db failed")
        return

    book = openpyxl.open(xlsx_env.XLSX_SCHEDULE_PATH, read_only=True)
    sheet = book.active

    schedule_strings = parse_current_schedule(book, sheet)

    book.close()
    if storage.save_data_schedule(schedule_strings) != invalid_returned_id:
        logging.info("update schedule db successfully")
    else:
        logging.error("update schedule db failed")


def getValueWithMergeLookup(sheet, cell):
    idx = cell.coordinate

    # for range_ in sheet.merged_cell_ranges:
    # 'merged_cell_ranges' has been deprecated
    # 'merged_cells.ranges' should be used instead
    for range_ in sheet.merged_cells.ranges:

        # merged_cells = list(openpyxl.utils.rows_from_range(range_))
        # 'rows_from_range' should take a 'str' type argument
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))

        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range

                # return sheet.cell(merged_cells[0][0]).value
                # You can just use 'sheet[<CELL ADDRESS>]' to take a cell
                # ex) sheet["A1"].value
                return sheet[merged_cells[0][0]].value

    # return sheet.cell(idx).value
    return sheet[idx].value


def get_merged_cell_value(sht, cell):
    """
    Check whether it is a merged cell and get the value of the corresponding row and column cell.
    If it is a merged cell, take the value of the cell in the upper left corner of the merged area as the value of the current cell; otherwise, directly return the value of the cell
    : param sht: current sheet object
    : param cell: current cell
    :return: value from merged cell
    """
    if isinstance(cell, openpyxl.cell.MergedCell):  # judge whether the cell is a merged cell
        for merged_range in sht.merged_cells.ranges:  # loop to find the merge range to which the cell belongs
            if cell.coordinate in merged_range:
                # Gets the cell in the upper left corner of the merge range and returns it as the value of the cell
                cell = sht.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


def parse_time_idx(book, sheet):
    dict_l = dict()
    for j in range(sheet.min_column - 1, sheet.max_column):
        for i in range(sheet.min_row - 1, sheet.max_row):
            if sheet[i][j].value is not None and sheet[i][j].value in xlsx_env.dday.keys():
                day = sheet[i][j].value
                time_idx = i
                dict_l[day] = []
                while sheet[time_idx][j + 1].value != xlsx_env.dtime[-1]:
                    if sheet[time_idx][j + 1].value in xlsx_env.dtime:
                        dict_l[day].append({str(sheet[time_idx][j + 1].value).replace(" ", ""): time_idx})
                    time_idx += 1
                dict_l[day].append({str(sheet[time_idx][j + 1].value).replace(" ", ""): time_idx})
                if sheet[time_idx][j + 1].value == xlsx_env.dtime[-1] and day == list(xlsx_env.dday.keys())[-1]:
                    return dict_l


def parse_group_idx(book, sheet):
    dict_l = dict()
    for j in range(sheet.min_column - 1, sheet.max_column):
        for i in range(sheet.min_row - 1, sheet.max_row):
            if sheet[i][j].value == list(xlsx_env.dday.keys())[0] or sheet[i][j].value == xlsx_env.dtime[0]:
                break
            if sheet[i][j].value is not None and sheet[i][j].value in xlsx_env.dgroup:
                group_idx = j
                while sheet[i][group_idx].value != xlsx_env.dgroup[-1]:
                    if sheet[i][group_idx].value in xlsx_env.dgroup:
                        dict_l[sheet[i][group_idx].value] = group_idx
                    group_idx += 1
                dict_l[sheet[i][group_idx].value] = group_idx
                return dict_l


def parse_current_schedule(book, sheet):
    logging.info("start parse group idx")
    group_idx = parse_group_idx(book, sheet)
    logging.info("start parse time idx")
    time_idx = parse_time_idx(book, sheet)

    book.close()
    # fix readonly flag
    book = openpyxl.open(xlsx_env.XLSX_SCHEDULE_PATH)
    sheet = book.active

    logging.info("start parse schedule strings")
    result = []
    result_str = ""
    format_day_str = "<b><u>{day_s}:</u></b>\n{time_s}\n{key_word}\n"
    format_time_str = "<b>{time_s}:</b> \t<i>{classwork}</i>\n"
    for group_key, group_value in group_idx.items():
        col = group_value
        day_str = f"<strong>Расписание для группы {group_key}:</strong>\n\n"
        for day_key, day_times in time_idx.items():
            time_str = ""
            for time in day_times:
                for time_key, time_row in time.items():
                    tmp = str(getValueWithMergeLookup(sheet, sheet[time_row][col]))
                    if tmp != None and len(tmp) > min_len_classwork:
                        time_str += format_time_str.format(time_s=time_key, classwork=tmp)
            if time_str != None and time_str != "":
                day_str += format_day_str.format(day_s=day_key, time_s=time_str, key_word=xlsx_env.KEY_WORD_NEXT_DAY)
        result.append({'group_key': group_key, 'day_str': day_str})
    book.close()
    return result


if __name__ == "__main__":
    update_schedule_db()
