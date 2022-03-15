import openpyxl
from openpyxl.cell import MergedCell

from xlsxparser import xlsx_env
from xlsxparser.db_utils import MongodbService

min_len_classwork = 10
book = openpyxl.open(xlsx_env.XLSX_SCHEDULE_PATH, read_only=True)

sheet = book.active

storage = MongodbService.get_instance()


def get_merged_cell_value(sht, cell):
    """
    Check whether it is a merged cell and get the value of the corresponding row and column cell.
    If it is a merged cell, take the value of the cell in the upper left corner of the merged area as the value of the current cell; otherwise, directly return the value of the cell
    : param sht: current sheet object
    : param cell: current cell
    :return: value from merged cell
    """
    if isinstance(cell, MergedCell):  # judge whether the cell is a merged cell
        for merged_range in sht.merged_cells.ranges:  # loop to find the merge range to which the cell belongs
            if cell.coordinate in merged_range:
                # Gets the cell in the upper left corner of the merge range and returns it as the value of the cell
                cell = sht.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


def parse_time_idx():
    dict_l = dict()
    for j in range(sheet.min_column - 1, sheet.max_column):
        for i in range(sheet.min_row - 1, sheet.max_row):
            if sheet[i][j].value is not None and sheet[i][j].value in xlsx_env.dday.keys():
                day = sheet[i][j].value
                time_idx = i
                dict_l[day] = []
                while sheet[time_idx][j + 1].value != xlsx_env.dtime[-1]:
                    if sheet[time_idx][j + 1].value in xlsx_env.dtime:
                        dict_l[day].append({str(sheet[time_idx][j + 1].value).replace(" ", ""): time_idx})
                    time_idx += 1
                dict_l[day].append({str(sheet[time_idx][j + 1].value).replace(" ", ""): time_idx})
                if sheet[time_idx][j + 1].value == xlsx_env.dtime[-1] and day == list(xlsx_env.dday.keys())[-1]:
                    return dict_l


def parse_group_idx():
    dict_l = dict()
    for j in range(sheet.min_column - 1, sheet.max_column):
        for i in range(sheet.min_row - 1, sheet.max_row):
            if sheet[i][j].value == list(xlsx_env.dday.keys())[0] or sheet[i][j].value == xlsx_env.dtime[0]:
                break
            if sheet[i][j].value is not None and sheet[i][j].value in xlsx_env.dgroup:
                group_idx = j
                while sheet[i][group_idx].value != xlsx_env.dgroup[-1]:
                    if sheet[i][group_idx].value in xlsx_env.dgroup:
                        dict_l[sheet[i][group_idx].value] = group_idx
                    group_idx += 1
                dict_l[sheet[i][group_idx].value] = group_idx
                return dict_l


def parse_current_schedule(group_idx=None, time_idx=None):
    group_idx = parse_group_idx()
    time_idx = parse_time_idx()

    result = []
    result_str = ""
    format_day_str = "**{day_s}:**\n{time_s}\n"
    format_time_str = "**{time_s}:** \t__{classwork}__\n"
    for group_key, group_value in group_idx.items():
        col = group_value
        day_str = f"**Расписание для группы {group_key}:**\n\n"
        for day_key, day_times in time_idx.items():
            time_str = ""
            for time in day_times:
                for time_key, time_row in time.items():
                    tmp = str(get_merged_cell_value(sheet, sheet[time_row][col]))
                    if tmp != None and len(tmp) > min_len_classwork:
                        time_str += format_time_str.format(time_s=time_key, classwork=tmp)
            if time_str != None and time_str != "":
                day_str += format_day_str.format(day_s=day_key, time_s=time_str)
        result.append({'group_key': group_key, 'day_str': day_str})
        print(day_str)
    return result


if __name__ == "__main__":
    group_idx = parse_group_idx()
    time_idx = parse_time_idx()
    # group_idx = {'09-101 (1)': 2, '09-101 (2)': 3, '09-102 (1)': 4, '09-102 (2)': 5, '09-103 (1)': 6, '09-103 (2)': 7, '09-111 (1)': 8, '09-111 (2)': 9, '09-112 (1)': 10, '09-112 (2)': 11, '09-113 (1)': 12, '09-113 (2)': 13, '09-121 (1)': 14, '09-121 (2)': 15, '09-122 (1)': 16, '09-122 (2)': 17, '09-131 (1)': 18, '09-131 (2)': 19, '09-132 (1)': 20, '09-132 (2)': 21, '09-141 (1)': 22, '09-141 (2)': 23, '09-142 (1)': 24, '09-142 (2)': 25, '09-151 (1)': 26, '09-151 (2)': 27, '09-152 (1)': 28, '09-152 (2)': 29, '09-153 (1)': 30, '09-153 (2)': 31, '09-161 (1)': 32, '09-161 (2)': 33, '09-162 (1)': 34, '09-162 (2)': 35, '09-163 (1)': 36, '09-163 (2)': 37, '09-001 (1)': 40, '09-001 (2)': 41, '09-002 (1)': 42, '09-002 (2)': 43, '09-011 (1)': 44, '09-011 (2)': 45, '09-012 (1)': 46, '09-012 (2)': 47, '09-013 (1)': 48, '09-013 (2)': 49, '09-021 (1)': 50, '09-021 (2)': 51, '09-022 (1)': 52, '09-022 (2)': 53, '09-031 (1)': 54, '09-031 (2)': 55, '09-032 (1)': 56, '09-032 (2)': 57, '09-033 (1)': 58, '09-033 (2)': 59, '09-034 (1)': 60, '09-034 (2)': 61, '09-041 (1)': 62, '09-041 (2)': 63, '09-042 (1)': 64, '09-042 (2)': 65, '09-051 (1)': 66, '09-051 (2)': 67, '09-052 (1)': 68, '09-052 (2)': 69, '09-053 (1)': 70, '09-053 (2)': 71, '09-061 (1)': 72, '09-061 (2)': 73, '09-062 (1)': 74, '09-062 (2)': 75, '09-063 (1)': 76, '09-063 (2)': 77, '09-901 (1)': 80, '09-901 (2)': 81, '09-902 (1)': 82, '09-902 (2)': 83, '09-911 (1)': 84, '09-911 (2)': 85, '09-912 (1)': 86, '09-912 (2)': 87, '09-913 (1)': 88, '09-913 (2)': 89, '09-921 (1)': 90, '09-921 (2)': 91, '09-922 (1)': 92, '09-922 (2)': 93, '09-931 (1)': 94, '09-931 (2)': 95, '09-932 (1)': 96, '09-932 (2)': 97, '09-933 (1)': 98, '09-933 (2)': 99, '09-941 (1)': 100, '09-941 (2)': 101, '09-942 (1)': 102, '09-942 (2)': 103, '09-951 (1)': 104, '09-951 (2)': 105, '09-952 (1)': 106, '09-952 (2)': 107, '09-961 (1)': 108, '09-961 (2)': 109, '09-962 (1)': 110, '09-962 (2)': 111, '09-963 (1)': 112, '09-963 (2)': 113, '09-801 (1)': 116, '09-801 (2)': 117, '09-802 (1)': 118, '09-802 (2)': 119, '09-811 (1)': 120, '09-811 (2)': 121, '09-812 (1)': 122, '09-812 (2)': 123, '09-813 (1)': 124, '09-813 (2)': 125, '09-821 (1)': 126, '09-821 (2)': 127, '09-822 (1)': 128, '09-822 (2)': 129, '09-831 (1)': 130, '09-831 (2)': 131, '09-832 (1)': 132, '09-832 (2)': 133, '09-833 (1)': 134, '09-833 (2)': 135, '09-841 (1)': 136, '09-841 (2)': 137, '09-851 (1)': 138, '09-851 (2)': 139, '09-852 (1)': 140, '09-852 (2)': 141, '09-861 (1)': 142, '09-861 (2)': 143, '09-862 (1)': 144, '09-862 (2)': 145}
    # time_idx = {'понедельник': [{'8.30-10.00': 33}, {'10.10-11.40': 35}, {'11.50-13.20': 37}, {'14.00-15.30': 39}, {'15.40-17.10': 41}, {'17.50-19.20': 43}, {'19.30-21.00': 45}], 'вторник': [{'8.30-10.00': 48}, {'10.10-11.40': 50}, {'11.50-13.20': 52}, {'14.00-15.30': 54}, {'15.40-17.10': 56}, {'17.50-19.20': 58}, {'19.30-21.00': 60}], 'среда': [{'8.30-10.00': 63}, {'10.10-11.40': 65}, {'11.50-13.20': 67}, {'14.00-15.30': 69}, {'15.40-17.10': 71}, {'17.50-19.20': 73}, {'19.30-21.00': 75}], 'четверг': [{'8.30-10.00': 78}, {'10.10-11.40': 80}, {'11.50-13.20': 82}, {'14.00-15.30': 84}, {'15.40-17.10': 86}, {'17.50-19.20': 88}, {'19.30-21.00': 90}], 'пятница': [{'8.30-10.00': 93}, {'10.10-11.40': 95}, {'11.50-13.20': 97}, {'14.00-15.30': 99}, {'15.40-17.10': 101}, {'17.50-19.20': 103}, {'19.30-21.00': 105}], 'суббота': [{'8.30-10.00': 108}, {'10.10-11.40': 110}, {'11.50-13.20': 112}, {'14.00-15.30': 114}, {'15.40-17.10': 116}, {'17.50-19.20': 118}, {'19.30-21.00': 120}]}

    print(parse_current_schedule(group_idx, time_idx))
